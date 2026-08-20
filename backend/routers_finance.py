import io
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Response, UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from core import db, require_admin, require_super, audit, new_id, now_iso

router = APIRouter(tags=["keuangan"])

DRAFT, PENDING, PUBLISHED = "draft", "menunggu_persetujuan", "terbit"


class ItemBody(BaseModel):
    type: str = Field(pattern="^(masuk|keluar)$")
    description: str = Field(min_length=2, max_length=200)
    amount: int = Field(gt=0)
    date: Optional[str] = None
    receipt_file_id: Optional[str] = None
    receipt_public: bool = True


class ReportBody(BaseModel):
    title_id: str = Field(min_length=3, max_length=160)
    title_en: Optional[str] = None
    event_date: Optional[str] = None
    description_id: Optional[str] = None
    description_en: Optional[str] = None
    items: List[ItemBody] = []
    revision_reason: Optional[str] = None


def recompute(doc: dict) -> dict:
    total_in = sum(i["amount"] for i in doc.get("items", []) if i["type"] == "masuk")
    total_out = sum(i["amount"] for i in doc.get("items", []) if i["type"] == "keluar")
    doc["total_in"] = total_in
    doc["total_out"] = total_out
    doc["balance"] = total_in - total_out
    return doc


def strip_private(doc: dict) -> dict:
    for item in doc.get("items", []):
        if not item.get("receipt_public", True):
            item["receipt_file_id"] = None
            item["receipt_hidden"] = True
    return doc


@router.get("/finance")
async def list_reports():
    reports = await db.finance.find({"status": PUBLISHED},
                                    {"_id": 0, "items": 0}).sort("event_date", -1).to_list(200)
    return reports


EXCEL_HEADERS = ["jenis", "tanggal", "keterangan", "jumlah"]
EXCEL_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@router.get("/admin/finance-template.xlsx")
async def finance_template(user: dict = Depends(require_admin)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rincian"
    header_fill = PatternFill("solid", fgColor="1A2F24")
    for i, h in enumerate(["Jenis (masuk/keluar)", "Tanggal (YYYY-MM-DD)", "Keterangan",
                           "Jumlah (Rupiah, angka bulat)"], start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = 30
    ws.row_dimensions[1].height = 30
    contoh = [
        ["masuk", "2026-06-10", "Iuran peserta turnamen", 3600000],
        ["masuk", "2026-06-12", "Donasi warga RW 02", 1500000],
        ["keluar", "2026-06-14", "Pembelian bola dan net", 1850000],
        ["keluar", "2026-06-20", "Konsumsi panitia", 1240000],
    ]
    for row in contoh:
        ws.append(row)
    for r in range(2, 2 + len(contoh)):
        ws.cell(row=r, column=4).number_format = "#,##0"

    info = wb.create_sheet("Petunjuk")
    for i, line in enumerate([
        "Cara memakai template ini",
        "",
        "1. Isi satu baris untuk setiap penerimaan atau pengeluaran.",
        "2. Kolom Jenis hanya boleh berisi: masuk atau keluar.",
        "3. Kolom Tanggal memakai format YYYY-MM-DD, contoh 2026-06-10.",
        "4. Kolom Jumlah diisi angka rupiah bulat tanpa titik, koma, atau tulisan 'Rp'.",
        "5. Hapus baris contoh sebelum mengunggah.",
        "6. Unggah berkas ini di menu Keuangan -> Impor Excel. Nota tetap diunggah manual",
        "   melalui alat sensor agar NIK dan nomor HP tidak ikut tersebar.",
    ], start=1):
        c = info.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=13)
    info.column_dimensions["A"].width = 95

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(), media_type=EXCEL_MEDIA,
        headers={"Content-Disposition": 'attachment; filename="template-laporan-keuangan.xlsx"'})


def _parse_amount(value) -> int:
    if value is None or value == "":
        raise ValueError("jumlah kosong")
    if isinstance(value, (int, float)):
        return int(round(value))
    txt = str(value).lower().replace("rp", "").replace(" ", "").replace(".", "").replace(",", "")
    if not txt.isdigit():
        raise ValueError(f"jumlah '{value}' bukan angka")
    return int(txt)


def _parse_date(value) -> Optional[str]:
    if value is None or value == "":
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


@router.post("/admin/finance-parse-excel")
async def parse_finance_excel(file: UploadFile = File(...), user: dict = Depends(require_admin)):
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Gunakan berkas Excel .xlsx sesuai template")
    raw = await file.read()
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Ukuran berkas maksimal 5MB")
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Berkas Excel tidak dapat dibaca")
    ws = wb["Rincian"] if "Rincian" in wb.sheetnames else wb.worksheets[0]

    items, errors = [], []
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=4, values_only=True), start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        jenis, tanggal, keterangan, jumlah = (list(row) + [None] * 4)[:4]
        jenis = str(jenis or "").strip().lower()
        if jenis in ("in", "pemasukan", "penerimaan"):
            jenis = "masuk"
        if jenis in ("out", "pengeluaran"):
            jenis = "keluar"
        if jenis not in ("masuk", "keluar"):
            errors.append(f"Baris {idx}: jenis harus 'masuk' atau 'keluar'")
            continue
        if not str(keterangan or "").strip():
            errors.append(f"Baris {idx}: keterangan wajib diisi")
            continue
        try:
            amount = _parse_amount(jumlah)
        except ValueError as e:
            errors.append(f"Baris {idx}: {e}")
            continue
        if amount <= 0:
            errors.append(f"Baris {idx}: jumlah harus lebih dari 0")
            continue
        items.append({"type": jenis, "date": _parse_date(tanggal),
                      "description": str(keterangan).strip()[:200], "amount": amount,
                      "receipt_file_id": None, "receipt_public": True})

    if not items and errors:
        raise HTTPException(status_code=400, detail=" · ".join(errors[:5]))
    await audit(user, "impor_excel_keuangan", "", f"{len(items)} item, {len(errors)} error")
    return {"items": items, "errors": errors,
            "total_in": sum(i["amount"] for i in items if i["type"] == "masuk"),
            "total_out": sum(i["amount"] for i in items if i["type"] == "keluar")}


@router.get("/finance/summary/yearly")
async def yearly_summary():
    reports = await db.finance.find({"status": PUBLISHED}, {"_id": 0}).to_list(500)
    buckets = {}
    for r in reports:
        year = (r.get("event_date") or r.get("published_at") or "")[:4]
        if not year:
            continue
        b = buckets.setdefault(year, {"year": year, "total_in": 0, "total_out": 0, "reports": 0})
        b["total_in"] += r.get("total_in", 0)
        b["total_out"] += r.get("total_out", 0)
        b["reports"] += 1
    rows = sorted(buckets.values(), key=lambda b: b["year"])
    for b in rows:
        b["balance"] = b["total_in"] - b["total_out"]
    return {"years": rows,
            "grand_total_in": sum(b["total_in"] for b in rows),
            "grand_total_out": sum(b["total_out"] for b in rows),
            "grand_balance": sum(b["balance"] for b in rows)}


@router.get("/finance/{report_id}")
async def get_report(report_id: str):
    doc = await db.finance.find_one({"id": report_id, "status": PUBLISHED}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Laporan tidak ditemukan atau belum terbit")
    return strip_private(doc)


@router.get("/admin/finance")
async def admin_list_reports(user: dict = Depends(require_admin)):
    return await db.finance.find({}, {"_id": 0}).sort("updated_at", -1).to_list(300)


@router.get("/admin/finance/{report_id}")
async def admin_get_report(report_id: str, user: dict = Depends(require_admin)):
    doc = await db.finance.find_one({"id": report_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Laporan tidak ditemukan")
    return doc


@router.post("/admin/finance")
async def create_report(body: ReportBody, user: dict = Depends(require_admin)):
    doc = body.model_dump(exclude={"revision_reason"})
    doc.update({"id": new_id(), "status": DRAFT, "created_by": user["name"],
                "created_at": now_iso(), "updated_at": now_iso(),
                "published_at": None, "revisions": []})
    recompute(doc)
    await db.finance.insert_one(doc)
    await audit(user, "buat_laporan_keuangan", doc["id"], body.title_id)
    doc.pop("_id", None)
    return doc


@router.put("/admin/finance/{report_id}")
async def update_report(report_id: str, body: ReportBody, user: dict = Depends(require_admin)):
    doc = await db.finance.find_one({"id": report_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Laporan tidak ditemukan")
    if doc["status"] == PUBLISHED and not (body.revision_reason or "").strip():
        raise HTTPException(status_code=400,
                            detail="Alasan revisi wajib diisi untuk laporan yang sudah terbit")
    updates = body.model_dump(exclude={"revision_reason"})
    updates["updated_at"] = now_iso()
    recompute(updates)
    ops = {"$set": updates}
    if doc["status"] == PUBLISHED:
        await db.finance_versions.insert_one({
            "id": new_id(), "report_id": report_id, "snapshot":
                {k: v for k, v in doc.items() if k != "_id"}, "at": now_iso(),
            "by": user["name"], "reason": body.revision_reason})
        ops["$push"] = {"revisions": {"at": now_iso(), "by": user["name"],
                                      "reason": body.revision_reason.strip()}}
    await db.finance.update_one({"id": report_id}, ops)
    await audit(user, "ubah_laporan_keuangan", report_id, body.revision_reason or body.title_id)
    return await db.finance.find_one({"id": report_id}, {"_id": 0})


@router.delete("/admin/finance/{report_id}")
async def delete_report(report_id: str, user: dict = Depends(require_super)):
    await db.finance.delete_one({"id": report_id})
    await audit(user, "hapus_laporan_keuangan", report_id)
    return {"ok": True}


@router.post("/admin/finance/{report_id}/submit")
async def submit_report(report_id: str, user: dict = Depends(require_admin)):
    doc = await db.finance.find_one({"id": report_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Laporan tidak ditemukan")
    if not doc.get("items"):
        raise HTTPException(status_code=400, detail="Laporan kosong tidak dapat diajukan")
    await db.finance.update_one({"id": report_id}, {"$set": {
        "status": PENDING, "submitted_at": now_iso(), "submitted_by": user["name"],
        "updated_at": now_iso()}})
    await audit(user, "ajukan_laporan", report_id, doc["title_id"])
    return await db.finance.find_one({"id": report_id}, {"_id": 0})


@router.post("/admin/finance/{report_id}/approve")
async def approve_report(report_id: str, user: dict = Depends(require_super)):
    doc = await db.finance.find_one({"id": report_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Laporan tidak ditemukan")
    if doc["status"] != PENDING:
        raise HTTPException(status_code=400, detail="Hanya laporan yang diajukan dapat disetujui")
    await db.finance.update_one({"id": report_id}, {"$set": {
        "status": PUBLISHED, "published_at": now_iso(), "approved_by": user["name"],
        "updated_at": now_iso()}})
    await audit(user, "setujui_dan_terbitkan_laporan", report_id, doc["title_id"])
    return await db.finance.find_one({"id": report_id}, {"_id": 0})


class RejectBody(BaseModel):
    reason: str = Field(min_length=3, max_length=300)


@router.post("/admin/finance/{report_id}/reject")
async def reject_report(report_id: str, body: RejectBody, user: dict = Depends(require_super)):
    doc = await db.finance.find_one({"id": report_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Laporan tidak ditemukan")
    await db.finance.update_one({"id": report_id}, {"$set": {
        "status": DRAFT, "reject_reason": body.reason, "updated_at": now_iso()}})
    await audit(user, "tolak_laporan", report_id, body.reason)
    return await db.finance.find_one({"id": report_id}, {"_id": 0})


@router.post("/admin/finance/{report_id}/unpublish")
async def unpublish_report(report_id: str, body: RejectBody, user: dict = Depends(require_super)):
    doc = await db.finance.find_one({"id": report_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Laporan tidak ditemukan")
    await db.finance.update_one({"id": report_id}, {"$set": {
        "status": DRAFT, "unpublish_reason": body.reason, "updated_at": now_iso()}})
    await audit(user, "tarik_laporan", report_id, body.reason)
    return await db.finance.find_one({"id": report_id}, {"_id": 0})


class ReceiptVisibilityBody(BaseModel):
    receipt_public: bool


@router.put("/admin/finance/{report_id}/items/{item_index}/receipt")
async def set_receipt_visibility(report_id: str, item_index: int, body: ReceiptVisibilityBody,
                                 user: dict = Depends(require_super)):
    doc = await db.finance.find_one({"id": report_id})
    if not doc or item_index >= len(doc.get("items", [])):
        raise HTTPException(status_code=404, detail="Item tidak ditemukan")
    await db.finance.update_one({"id": report_id}, {"$set": {
        f"items.{item_index}.receipt_public": body.receipt_public, "updated_at": now_iso()}})
    await audit(user, "ubah_visibilitas_nota", report_id,
                f"item {item_index} -> {'publik' if body.receipt_public else 'disembunyikan'}")
    return await db.finance.find_one({"id": report_id}, {"_id": 0})


@router.get("/finance/{report_id}/versions")
async def report_versions(report_id: str):
    versions = await db.finance_versions.find({"report_id": report_id},
                                              {"_id": 0}).sort("at", -1).to_list(50)
    return versions
